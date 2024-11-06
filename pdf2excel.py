import numpy as np
import re
from openpyxl import Workbook, load_workbook
import os
from pdfminer.high_level import extract_text

# Funzione per pulire il testo, rimuovendo righe vuote, spazi e caratteri superflui
def clean_text(paragrafo, file_name):
    # Sostituisce il carattere di nuova riga tra parole con uno spazio
    paragrafo = re.sub(r'(?<=[a-zA-Z])\n|(?=\n[a-zA-Z])', ' ', paragrafo)
    paragrafo = re.sub(r'\n+', '', paragrafo)
    paragrafo = re.sub(r'\s+', ' ', paragrafo)  # Rimuove spazi multipli
    paragrafo = paragrafo.replace('- ', '')     # Rimuove trattini seguiti da spazio
    # Rimuove l'identificativo del file se presente
    file_name = re.sub('.pdf', '', file_name)
    paragrafo = re.sub(rf'1\s+{re.escape(file_name)}', '', paragrafo)
    return paragrafo

# Caricamento del file Excel
#                                  DA MODIFICARE
#                                         |
#                                         V
excel_path = '/Users/francesconuzzo/Downloads/PROCEEDINGS/Repository Upload Information_IFASD_2019 copia.xlsx'  # Modifica con il percorso del file
wb = load_workbook(excel_path)
ws = wb.active

# Definizione dei limiti delle righe da processare (righe di excel)
raw_begin = 9

#     DA MODIFICARE
#         |
#         V
raw_end = 139
titles = ''

# Identificazione dell'indice di partenza
indice = ws[f'E{raw_begin - 1}'].value + 1 if isinstance(ws[f'E{raw_begin - 1}'].value, int) else raw_begin - 8

# Iterazione sulle righe per estrarre e processare le informazioni
for row in range(raw_begin, raw_end):
    if indice >= raw_end-1:
        break

    # Flag per verificare la presenza delle informazioni necessarie
    flag_authors = flag_conf_numb = flag_abstract = flag_file_name = flag_keywords = flag_category = flag_online = True
    flag_title = ws[f'B{row}'].value is None
    if ws[f'C{row}'].value is None:
        flag_authors = False
    if ws[f'E{row}'].value is None:
        flag_conf_numb = False
    if ws[f'F{row}'].value is None:
        flag_file_name = False
    if ws[f'G{row}'].value is None:
        flag_abstract = False
    if ws[f'H{row}'].value is None:
        flag_category = False
    if ws[f'I{row}'].value is None:
        flag_keywords = False
    if ws[f'J{row}'].value is None:
        flag_online = False

    # Se qualche flag è False, cerca il file corrispondente
    if not (flag_title and flag_authors and flag_conf_numb and flag_file_name and flag_category and flag_online and flag_keywords and flag_abstract):
        is_file_flag = False
        while not is_file_flag and indice <= 208:
            #!!!!!!       DA MODIFICARE
            #                  |
            #                  V
            IFASD_name = f"IFASD-2019-{indice:03}.pdf"  # Genera il nome file con indice a 3 cifre
            
            #!!!!!!!!!!!!!!!!!!!!!!!!!         DA MODIFICARE
            #                                         |
            #                                         V
            file_path = os.path.join('/Users/francesconuzzo/Downloads/PROCEEDINGS', IFASD_name)  # Modifica con il percorso cartella
            if os.path.isfile(file_path):
                is_file_flag = True
            else:
                indice += 1

        if not flag_conf_numb:
            ws[f'E{row}'].value = indice
        if is_file_flag:
            indice += 1

        # Apertura e lettura del testo dal PDF solo una volta per riga
        if not is_file_flag:
            continue

        text = extract_text(file_path, page_numbers=[0, 1])
        paragrafo = clean_text(text, IFASD_name)

        # Estrazione del titolo con tutti i pattern specificati
        if flag_title or not flag_authors:
            title_patterns = [
                r'USA\s*[\r\n]*(.*?)\s*(?:\n\s*\n|$)',
                r'\s*(.*?)(?:\r?\n\r?\n)',                 # Pattern generico per il titolo
                r'\s*Title\s*:\s*(.*?)(?:\r?\n\r?\n)',     # Pattern per "Title :"
                r'\s*Paper Title\s*:\s*(.*?)(?:\r?\n\r?\n)',  # Pattern per "Paper Title :"
                r'\s*(.*?)\s*Authors',                     # Pattern che termina prima di "Authors"
                r'\s*(.*?)\s*Introduction',                # Pattern che termina prima di "Introduction"
            ]
            for pattern in title_patterns:
                match = re.search(pattern, text, re.DOTALL)
                if match:
                    paper_title_raw = match.group(1).strip()
                    paper_title = clean_text(paper_title_raw, IFASD_name)
                    if flag_title: ws[f'B{row}'].value = paper_title
                    print(paper_title + '\n')
                    break

        # Estrazione degli autori con pattern multipli
        if not flag_authors:
            # Pattern per trovare gli autori dopo il titolo
            authors_patterns = [
                rf'{re.escape(paper_title_raw)}\n*(.+?)(?=\n{{2,}})',
                paper_title_raw + r'\n*([^\n]+)',
                paper_title_raw + r'\n*(.+?)(?=\n{2,})',
                re.escape(paper_title_raw) + r'\n*([^\n]+)',
                re.escape(paper_title_raw) + r'\n*(.+?)(?=\n{2,})',
                
                r'{re.escape(paper_title_raw)}*\n*(.+?)(?=\n{{2,}})',  # Cerca gli autori dopo il titolo
                r'\s*Authors\s*[:\-]?\s*(.*?)\s*(?=\n{2,})',      # Autori dopo "Authors:"
                r'\s*(.*?)\s*1\s+Introduction',                    # Autori prima di "Introduction"
            ]
            
            for pattern in authors_patterns:
                match = re.findall(pattern, text, re.DOTALL)
                if match:
                    # Rimuovi numeri, trattini e caratteri non necessari
                    paper_authors = re.sub(r'\d+', '', match[0])  # Rimuove numeri
                    paper_authors = re.sub(r'\n+', ' ', paper_authors)  # Rimuove eventuali ritorni a capo extra
                    paper_authors = paper_authors.strip()  # Rimuove spazi extra iniziali e finali
                    print(paper_authors + '\n\n')
                    # Salva gli autori nel foglio Excel
                    ws[f'C{row}'].value = clean_text(paper_authors, IFASD_name)
                    break

        # Inserisce il nome file
        if not flag_file_name:
            ws[f'F{row}'].value = IFASD_name

        # Estrazione dell'abstract con tutti i pattern richiesti
        if not flag_abstract:
            abstract_patterns = [
                r'Abstract:\s*(.*?)[\s*\n]?\d\s*INTRODUCTION',
                r'Abstract:\s*(.*?)[\s*\n]?\s*List of Symbols',
                r'Abstract:\s*(.*?)[\s*\n]?\s*Introduction',
                r'Abstract:\s*(.*?)[\s*\n]?\s*Notice to Readers',
                r'Abstract:?\s*(.*?)\s*\n*\s*INTRODUCTION',
                r'Abstract.\s*(.*?)[\s*\n]?[\d]\s*INTRODUCTION'
            ]
            for pattern in abstract_patterns:
                match = re.search(pattern, paragrafo, re.DOTALL)
                if match:
                    ws[f'G{row}'].value = match.group(1).strip()
                    break

        # Estrazione delle parole chiave con tutti i pattern specificati
        if not flag_keywords:
            keywords_patterns = [
                r'Keywords:\s*(.*?)\s*[\n\s]*Abstract\s*:',
                r'Keywords:\s*(.*?)[\s*\n]?\s*INTRODUCTION',
                r'Keywords:\s*(.*?)\s*1\s+INTRODUCTION',
                r'Key words:\s*(.*?)[\s*\n]?\s*Abstract'
            ]
            for pattern in keywords_patterns:
                match = re.search(pattern, paragrafo, re.DOTALL)
                if match:
                    ws[f'I{row}'].value = clean_text(match.group(1), IFASD_name)
                    break

        # Salva il titolo per il file di output
        title = ws[f'B{row}'].value
        if title:
            title = re.sub(r'\s+', ' ', title).strip()  # Rimuove spazi e righe extra
            titles += title + '\n'

# Scrittura dei titoli in un file di testo
with open('titles.txt', "w") as file_titles:
    file_titles.write(titles)

# Salvataggio del workbook Excel
wb.save(excel_path)
